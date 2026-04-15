"""
Import Master Dex challenge entries from the "template - master dex challenge" sheet.
Columns (0-based index): Box=0, Row=1, Slot=2, National#=3, Name=4, ImageURL=5, Games=6, Notes=7, Caught=8. Ignore last column.
Sections are detected by header rows containing: Living Dex, Gender variants, Stars, Shiny Living, Shiny Gender.
"""
import re
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from tracker.models import DexEntry
from tracker.bulbapedia_urls import bulbapedia_species_url
from tracker.db_snapshots import create_snapshot


SHEET_NAME = 'template - master dex challenge'
# Column indices (0-based) in the sheet: Box, Row, Slot, National#, Name, Image URL, Games, Notes, Caught
COL_BOX, COL_ROW, COL_SLOT = 0, 1, 2
COL_NATIONAL, COL_NAME, COL_IMAGE_URL = 4, 5, 6
COL_GAMES, COL_NOTES, COL_CAUGHT = 8, 9, 11
# Section header keywords (row is assigned to section when we see this in the row or in a dedicated header)
SECTION_KEYWORDS = [
    ('living_dex', ['living dex', 'living dex challenge']),
    ('gender_forms', ['gender variants', 'gender forms', 'forms']),
    ('stars', ['stars']),
    ('shiny_living_dex', ['shiny living', 'shiny living dex']),
    ('shiny_gender_forms', ['shiny gender', 'shiny forms']),
]


def _cell_value(ws, row, col):
    """Get cell value 1-based."""
    try:
        v = ws.cell(row=row, column=col + 1).value
        return v if v is None else str(v).strip()
    except Exception:
        return None


def _section_from_row(ws, row, num_cols):
    """Detect if this row is a section header; return section key or None."""
    for section_key, keywords in SECTION_KEYWORDS:
        for c in range(num_cols):
            val = _cell_value(ws, row, c)
            if val:
                val_lower = val.lower()
                for kw in keywords:
                    if kw in val_lower:
                        return section_key
    return None


def _parse_int(val, default=None):
    if val is None or val == '':
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _parse_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ('1', 'true', 'yes', 'x', 'y'):
        return True
    if s in ('0', 'false', 'no', ''):
        return False
    return False


def _is_data_row(ws, row, num_cols):
    """Row has at least box, row, slot as numbers and a name."""
    box = _parse_int(_cell_value(ws, row, COL_BOX))
    r = _parse_int(_cell_value(ws, row, COL_ROW))
    slot = _parse_int(_cell_value(ws, row, COL_SLOT))
    name = _cell_value(ws, row, COL_NAME)
    return box is not None and r is not None and slot is not None and name


def import_sheet(ws, stdout):
    num_cols = max(COL_CAUGHT + 1, 10)
    current_section = 'living_dex'
    sort_order = 0
    created = updated = 0
    for row in range(1, ws.max_row + 1):
        section = _section_from_row(ws, row, num_cols)
        if section:
            current_section = section
            sort_order = 0
            continue
        if not _is_data_row(ws, row, num_cols):
            continue
        sort_order += 1
        box = _parse_int(_cell_value(ws, row, COL_BOX), 0)
        r = _parse_int(_cell_value(ws, row, COL_ROW), 0)
        slot = _parse_int(_cell_value(ws, row, COL_SLOT), 0)
        national = _parse_int(_cell_value(ws, row, COL_NATIONAL), 0)
        name = (_cell_value(ws, row, COL_NAME) or '').strip() or f'#{national}'
        image_url = (_cell_value(ws, row, COL_IMAGE_URL) or '').strip()
        games = (_cell_value(ws, row, COL_GAMES) or '').strip()
        notes = (_cell_value(ws, row, COL_NOTES) or '').strip()
        caught = _parse_bool(_cell_value(ws, row, COL_CAUGHT))
        bulbapedia_url = (bulbapedia_species_url(name) or '')[:512]
        # Idempotent: match by (box, row, slot, section)
        entry, was_created = DexEntry.objects.update_or_create(
            box=box, row=r, slot=slot, section=current_section,
            defaults={
                'national_dex_number': national,
                'name': name,
                'bulbapedia_url': bulbapedia_url,
                'image_url': image_url[:512] if image_url else '',
                'games': games[:256] if games else '',
                'notes': notes,
                'caught': caught,
                'sort_order': sort_order,
            }
        )
        if was_created:
            created += 1
        else:
            updated += 1
    return created, updated


class Command(BaseCommand):
    help = 'Import Master Dex entries from Excel (sheet: template - master dex challenge)'

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='?', type=str, help='Path to .xlsx file')
        parser.add_argument('--dry-run', action='store_true', help='Only report what would be imported')

    def handle(self, *args, **options):
        path = options.get('path')
        if not path:
            self.stdout.write(self.style.ERROR('Usage: python manage.py import_master_dex_excel <path-to.xlsx>'))
            return
        dry_run = options.get('dry_run', False)
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to load workbook: {e}'))
            self.stdout.write('If the file has encoding issues, try saving the sheet as a new Excel file or CSV.')
            return
        sheet = None
        for name in wb.sheetnames:
            if name.strip().lower() == SHEET_NAME.strip().lower():
                sheet = wb[name]
                break
        if not sheet:
            self.stdout.write(self.style.ERROR(f'Sheet "{SHEET_NAME}" not found. Available: {wb.sheetnames}'))
            return
        if dry_run:
            self.stdout.write(f'[DRY RUN] Would import from sheet "{sheet.title}", max_row={sheet.max_row}')
            return
        snapshot = create_snapshot(label="before_import")
        self.stdout.write(f"Snapshot created: {snapshot}")
        created, updated = import_sheet(sheet, self.stdout)
        self.stdout.write(self.style.SUCCESS(f'Created {created}, updated {updated} entries.'))
